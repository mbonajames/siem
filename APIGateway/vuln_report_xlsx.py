"""
Technical report generator — produces an Excel workbook with:
  • Summary   sheet: totals side-by-side for internal vs external
  • Internal  sheet: all internal findings, auto-filter, frozen header, severity-colored rows
  • External  sheet: same layout for external findings
"""
import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Severity palette ──────────────────────────────────────────────────────────
_FILL = {
    "critical": PatternFill("solid", fgColor="FFB3B3"),
    "high":     PatternFill("solid", fgColor="FFD9B3"),
    "medium":   PatternFill("solid", fgColor="FFF5B3"),
    "low":      PatternFill("solid", fgColor="B3D9FF"),
    "info":     PatternFill("solid", fgColor="E8E8E8"),
}
_SEV_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}

_HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_TITLE_FONT    = Font(name="Calibri", bold=True, size=14, color="1F3864")
_LABEL_FONT    = Font(name="Calibri", bold=True, size=10)
_CELL_FONT     = Font(name="Calibri", size=9)
_CELL_ALIGN    = Alignment(wrap_text=True, vertical="top")
_THIN          = Side(style="thin", color="CCCCCC")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER        = Alignment(horizontal="center", vertical="center")

_FINDINGS_HEADERS = [
    "Severity", "Affected Host(s)", "Port", "Protocol",
    "Plugin ID", "Vulnerability Name",
    "CVE", "CVSS v3", "CVSS v2", "Risk Factor",
    "Description", "Solution",
]

_COL_WIDTHS = [12, 28, 8, 10, 10, 40, 22, 9, 9, 12, 60, 60]


# ── helpers ───────────────────────────────────────────────────────────────────

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws, row: int, headers: list):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = _CENTER


def _dedup_findings(scan_list: list) -> list:
    """Deduplicate by plugin_id across all scans; merge affected hosts into one row."""
    seen: dict = {}   # plugin_key → index in result
    result: list = []
    for scan in scan_list:
        for f in scan.get("findings", []):
            key = f.get("plugin_id") or f.get("plugin_name", "unknown")
            if key not in seen:
                entry = {**f, "_hosts": set()}
                h = f.get("host", "")
                if h:
                    entry["_hosts"].add(h)
                seen[key] = len(result)
                result.append(entry)
            else:
                h = f.get("host", "")
                if h:
                    result[seen[key]]["_hosts"].add(h)
    for f in result:
        hosts = sorted(f.pop("_hosts", set()))
        f["host"] = ", ".join(hosts) if hosts else "—"
    return sorted(result, key=lambda f: _SEV_ORDER.get(f.get("severity", "info"), 4))


def _sev_counts(deduped: list) -> dict:
    """Count unique findings by severity from a deduplicated list."""
    counts = {"critical": 0, "high": 0, "medium": 0, "low": 0, "info": 0}
    for f in deduped:
        sev = f.get("severity", "info")
        counts[sev] = counts.get(sev, 0) + 1
    counts["total_findings"] = len(deduped)
    return counts


def _write_findings_sheet(ws, findings: list, sheet_label: str):
    ws.title = sheet_label

    # ── title row ──
    ws.merge_cells("A1:L1")
    tc = ws["A1"]
    tc.value     = f"{sheet_label} Findings — {len(findings)} total"
    tc.font      = _TITLE_FONT
    tc.alignment = _CENTER

    # ── header row ──
    _write_header_row(ws, 2, _FINDINGS_HEADERS)
    ws.freeze_panes = "A3"

    # ── data rows ──
    for r, f in enumerate(findings, 3):
        sev  = f.get("severity", "info")
        fill = _FILL.get(sev, _FILL["info"])
        cves = ", ".join(f.get("cve") or [])

        row_vals = [
            sev.upper(),
            f.get("host", ""),
            f.get("port", ""),
            f.get("protocol", ""),
            f.get("plugin_id", ""),
            f.get("plugin_name", ""),
            cves,
            f.get("cvss3_base") or "",
            f.get("cvss_base")  or "",
            f.get("risk_factor", ""),
            f.get("description", "")[:500],   # cap for cell size
            f.get("solution", "")[:500],
        ]
        for c, val in enumerate(row_vals, 1):
            cell          = ws.cell(row=r, column=c, value=val)
            cell.fill     = fill
            cell.font     = _CELL_FONT
            cell.border   = _BORDER
            cell.alignment = _CELL_ALIGN

    # ── auto filter ──
    if findings:
        last_row = 2 + len(findings)
        ws.auto_filter.ref = f"A2:{get_column_letter(len(_FINDINGS_HEADERS))}{last_row}"

    _set_col_widths(ws, _COL_WIDTHS)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20


def _write_summary_sheet(ws, mfi: str, year: int, quarter: str,
                          internal_scans: list, external_scans: list,
                          int_deduped: list, ext_deduped: list):
    ws.title = "Summary"

    # ── report title ──
    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value     = f"Vulnerability Assessment — {mfi}  |  {quarter} {year}"
    tc.font      = Font(name="Calibri", bold=True, size=16, color="1F3864")
    tc.alignment = _CENTER
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"].value     = "Technical Summary Report (unique vulnerabilities by name)"
    ws["A2"].font      = Font(name="Calibri", size=11, color="666666")
    ws["A2"].alignment = _CENTER

    # ── severity totals — computed from deduplicated findings ──
    int_counts  = _sev_counts(int_deduped)
    ext_counts  = _sev_counts(ext_deduped)
    all_deduped = _dedup_findings(internal_scans + external_scans)
    all_counts  = _sev_counts(all_deduped)

    int_hosts = sum(len(set(h["ip"] for h in s.get("hosts", []))) for s in internal_scans)
    ext_hosts = sum(len(set(h["ip"] for h in s.get("hosts", []))) for s in external_scans)

    headers = ["", "Critical", "High", "Medium", "Low", "Info", "Unique Vulns", "Total Hosts"]
    rows = [
        ["Internal",
         int_counts["critical"], int_counts["high"], int_counts["medium"],
         int_counts["low"],      int_counts["info"],  int_counts["total_findings"], int_hosts],
        ["External",
         ext_counts["critical"], ext_counts["high"], ext_counts["medium"],
         ext_counts["low"],      ext_counts["info"],  ext_counts["total_findings"], ext_hosts],
        ["Combined",
         all_counts["critical"], all_counts["high"], all_counts["medium"],
         all_counts["low"],      all_counts["info"],  all_counts["total_findings"],
         int_hosts + ext_hosts],
    ]

    start_row = 4
    # header row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.border    = _BORDER
        cell.alignment = _CENTER

    sev_fills = [None,
                 PatternFill("solid", fgColor="FFB3B3"),  # critical
                 PatternFill("solid", fgColor="FFD9B3"),  # high
                 PatternFill("solid", fgColor="FFF5B3"),  # medium
                 PatternFill("solid", fgColor="B3D9FF"),  # low
                 PatternFill("solid", fgColor="E8E8E8"),  # info
                 None, None]

    for r_offset, row in enumerate(rows, 1):
        row_num = start_row + r_offset
        for c, val in enumerate(row, 1):
            cell          = ws.cell(row=row_num, column=c, value=val)
            cell.font     = Font(name="Calibri", bold=(c == 1 or r_offset == 3), size=10)
            cell.border   = _BORDER
            cell.alignment = _CENTER
            if sev_fills[c - 1]:
                cell.fill = sev_fills[c - 1]
            if r_offset == 3:
                cell.fill = PatternFill("solid", fgColor="E8F0FE")

    ws.column_dimensions["A"].width = 14
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16

    # ── file list ──
    ws.cell(row=start_row + 5, column=1, value="Source files:").font = _LABEL_FONT
    all_files = [(s.get("scan_type", ""), s.get("filename", "")) for s in internal_scans + external_scans]
    for i, (st, fn) in enumerate(all_files):
        ws.cell(row=start_row + 6 + i, column=1, value=f"  [{st.upper()}]  {fn}").font = _CELL_FONT


# ── public API ────────────────────────────────────────────────────────────────

def generate_technical_xlsx(mfi: str, year: int, quarter: str,
                             internal_scans: list, external_scans: list) -> bytes:
    int_deduped = _dedup_findings(internal_scans)
    ext_deduped = _dedup_findings(external_scans)

    wb = Workbook()
    ws_summary = wb.active
    _write_summary_sheet(ws_summary, mfi, year, quarter,
                         internal_scans, external_scans,
                         int_deduped, ext_deduped)

    ws_int = wb.create_sheet()
    _write_findings_sheet(ws_int, int_deduped, "Internal")

    ws_ext = wb.create_sheet()
    _write_findings_sheet(ws_ext, ext_deduped, "External")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
